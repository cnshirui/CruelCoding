#!/usr/bin/env python3
"""Build the web snapshot from the legacy workbook and membership files."""
import json
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "lc-score-board/generateEXCEL/index.xlsx"
IDS = ROOT / "lc-score-board/getRank/id.in"
MEMBERS = ROOT / "lc-score-board/generateEXCEL/Data/Members/In.txt"
OUTPUT = ROOT / "data/leaderboard.json"
norm = lambda value: str(value).strip().casefold()

cruel_ids = {norm(line): line.strip() for line in IDS.read_text().splitlines() if line.strip()}
membership = {}
for line in MEMBERS.read_text().splitlines():
    parts = line.split()
    if len(parts) >= 2:
        membership[norm(parts[0])] = {"date": datetime.strptime(parts[1], "%m/%d/%Y").date().isoformat(), "subgroup": parts[2] if len(parts) >= 3 else None}

sheet = load_workbook(WORKBOOK, data_only=True).active
contests = [(column, int(sheet.cell(9, column).value), int(sheet.cell(10, column).value or 0)) for column in range(6, sheet.max_column + 1, 2) if sheet.cell(9, column).value is not None]
rows = []
for row in range(12, sheet.max_row + 1):
    username = sheet.cell(row, 2).value
    if not username or not isinstance(sheet.cell(row, 1).value, (int, float)):
        continue
    key = norm(username)
    if key not in cruel_ids or key not in membership:
        raise ValueError(f"No CruelID/CruelDate match for workbook user {username!r}")
    history = []
    for column, contest, participants in contests:
        rank = int(sheet.cell(row, column).value or -1)
        history.append({"contest": contest, "participants": participants, "rank": rank if rank > 0 else None, "score": float(sheet.cell(row, column + 1).value or 0)})
    rows.append({"cruel_id": cruel_ids[key], "cruel_date": membership[key]["date"], "subgroup": membership[key]["subgroup"], "days": int(sheet.cell(row, 3).value or 0), "rating": int(sheet.cell(row, 4).value or -1), "score": float(sheet.cell(row, 5).value or 0), "contests": history})

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n")
print(f"Wrote {len(rows)} members and {len(contests)} contests to {OUTPUT.relative_to(ROOT)}")
